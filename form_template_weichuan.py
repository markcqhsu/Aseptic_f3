"""
Build 味全產品代工明細表 Excel from accumulated session data.

Column layout (1-indexed):
  A=1  日期
  B=2  所別/備註
  C=3  黑咖啡_全台   D=4  黑咖啡_4入
  E=5  拿鐵_全台     F=6  拿鐵_4入
  G=7  (reserved)    H=8  (reserved)
  I=9  全台板 T-11   (pallet formula =SUM(C:H)/80)

Data rows: 5-27 (23 rows)
Subtotal row: 28
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WEICHUAN_PRODUCT_CODES = ["黑咖啡_全台", "黑咖啡_4入", "拿鐵_全台", "拿鐵_4入"]

WEICHUAN_PRODUCT_LABELS = {
    "黑咖啡_全台": "貝納頌\n茶韻黑咖啡340ML\n全台",
    "黑咖啡_4入":  "貝納頌\n茶韻黑咖啡340ML\n全台4入",
    "拿鐵_全台":   "貝納頌\n茶韻拿鐵340ML\n全台",
    "拿鐵_4入":    "貝納頌\n茶韻拿鐵340ML\n全台4入",
}

CODE_TO_COL = {
    "黑咖啡_全台": 3,
    "黑咖啡_4入":  4,
    "拿鐵_全台":   5,
    "拿鐵_4入":    6,
}

COL_DATE   = 1   # A
COL_REMARK = 2   # B
COL_C      = 3   # first product
COL_F      = 6   # last product
COL_G      = 7   # reserved
COL_H      = 8   # reserved
COL_PAL    = 9   # I

DATA_ROW_START = 5
DATA_ROW_END   = 27
TOTAL_ROW      = 28

FONT_TITLE    = Font(name="微軟正黑體", size=16, bold=True)
FONT_SUBTITLE = Font(name="微軟正黑體", size=14, bold=True)
FONT_SECTION  = Font(name="微軟正黑體", size=14, bold=True)
FONT_REMARK   = Font(name="微軟正黑體", size=16, bold=True)
FONT_HEADER   = Font(name="微軟正黑體", size=12, bold=False)
FONT_SMALL    = Font(name="微軟正黑體", size=10, bold=False)
FONT_NORMAL   = Font(name="微軟正黑體", size=12, bold=False)
FONT_BOLD     = Font(name="微軟正黑體", size=12, bold=True)

FILL_BLUE = PatternFill("solid", fgColor="BDD7EE")

ALIGN_CENTER      = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT_WRAP   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_RIGHT_TOP   = Alignment(horizontal="right",  vertical="top",    wrap_text=True)
ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin   = Side(style="thin")
medium = Side(style="medium")
double = Side(style="double")


def build_weichuan_workbook(rows: list) -> bytes:
    """
    rows: list of dicts
      {
        "date":      "2026/04/23",
        "warehouse": "味全斗六廠115.04.23-02",
        "items":     [{"code": "拿鐵_全台", "qty": 2480}, ...]
      }
    Returns xlsx bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "代工明細表"

    _setup_columns(ws)
    _build_header(ws)
    _fill_data(ws, rows)
    _build_subtotal_row(ws)
    ws.freeze_panes = ws.cell(DATA_ROW_START, COL_C)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _setup_columns(ws):
    ws.column_dimensions["A"].width = 12.5
    ws.column_dimensions["B"].width = 30.0
    ws.column_dimensions["C"].width = 14.2
    ws.column_dimensions["D"].width = 14.7
    ws.column_dimensions["E"].width = 13.3
    ws.column_dimensions["F"].width = 12.7
    ws.column_dimensions["G"].width = 0.5
    ws.column_dimensions["H"].width = 10.2
    ws.column_dimensions["I"].width = 13.0

    ws.row_dimensions[1].height = 23.25
    ws.row_dimensions[2].height = 20.25
    ws.row_dimensions[3].height = 21.75
    ws.row_dimensions[4].height = 56.25


def _build_header(ws):
    # Row 1: Company name (merged A1:I1)
    ws.merge_cells("A1:I1")
    c = ws.cell(row=1, column=COL_DATE)
    c.value     = "宏全國際股份有限公司 無菌三廠"
    c.font      = FONT_TITLE
    c.alignment = ALIGN_CENTER

    # Row 2: Sheet title (merged A2:I2)
    ws.merge_cells("A2:I2")
    c = ws.cell(row=2, column=COL_DATE)
    c.value     = "味全產品代工明細表"
    c.font      = FONT_SUBTITLE
    c.alignment = ALIGN_CENTER

    # Row 3: 品項(A3), 出入處(B3), 出庫(C3:I3 merged)
    c = ws.cell(row=3, column=COL_DATE)
    c.value     = "品 項"
    c.font      = FONT_HEADER
    c.alignment = ALIGN_RIGHT_TOP
    c.border    = Border(top=medium, left=medium, right=medium)

    c = ws.cell(row=3, column=COL_REMARK)
    c.value     = "出入處"
    c.font      = FONT_SECTION
    c.alignment = ALIGN_CENTER
    c.fill      = FILL_BLUE
    c.border    = Border(top=medium, bottom=medium)

    for col in range(COL_C, COL_PAL + 1):
        ws.cell(row=3, column=col).border = Border(top=medium, bottom=medium)
    ws.cell(row=3, column=COL_C).border   = Border(top=medium, bottom=medium, left=medium)
    ws.cell(row=3, column=COL_PAL).border = Border(top=medium, bottom=medium, right=medium)
    ws.merge_cells(start_row=3, start_column=COL_C, end_row=3, end_column=COL_PAL)
    c = ws.cell(row=3, column=COL_C)
    c.value     = "出  庫"
    c.font      = FONT_SECTION
    c.alignment = ALIGN_CENTER
    c.fill      = FILL_BLUE

    # Row 4: Column labels
    c = ws.cell(row=4, column=COL_DATE)
    c.value     = "日 期"
    c.font      = FONT_HEADER
    c.alignment = ALIGN_LEFT_WRAP
    c.border    = Border(bottom=medium, left=medium)

    c = ws.cell(row=4, column=COL_REMARK)
    c.value     = "所別/備註"
    c.font      = FONT_REMARK
    c.alignment = ALIGN_CENTER
    c.fill      = FILL_BLUE
    c.border    = Border(bottom=medium, right=double)

    for i, code in enumerate(WEICHUAN_PRODUCT_CODES):
        col = COL_C + i
        c = ws.cell(row=4, column=col)
        c.value     = WEICHUAN_PRODUCT_LABELS[code]
        c.font      = FONT_SMALL
        c.alignment = ALIGN_CENTER_WRAP
        c.border    = Border(bottom=thin, right=thin)

    ws.cell(row=4, column=COL_G).font      = FONT_SMALL
    ws.cell(row=4, column=COL_G).alignment = ALIGN_CENTER_WRAP
    ws.cell(row=4, column=COL_G).border    = Border(bottom=thin, left=thin, right=thin)

    ws.cell(row=4, column=COL_H).font      = FONT_SMALL
    ws.cell(row=4, column=COL_H).alignment = ALIGN_CENTER_WRAP
    ws.cell(row=4, column=COL_H).border    = Border(bottom=thin, left=thin)

    c = ws.cell(row=4, column=COL_PAL)
    c.value     = "全台板\nT-11"
    c.font      = FONT_HEADER
    c.alignment = ALIGN_CENTER_WRAP
    c.fill      = FILL_BLUE
    c.border    = Border(bottom=thin, left=thin, right=thin)


def _fill_data(ws, rows: list):
    for row_idx, entry in enumerate(rows):
        r = DATA_ROW_START + row_idx
        if r > DATA_ROW_END:
            break

        _style_data_row(ws, r)

        ws.cell(row=r, column=COL_DATE).value   = entry.get("date", "")
        ws.cell(row=r, column=COL_REMARK).value = entry.get("warehouse", "")

        for item in entry.get("items", []):
            col = CODE_TO_COL.get(item["code"])
            if col:
                existing = ws.cell(row=r, column=col).value or 0
                ws.cell(row=r, column=col).value = existing + item["qty"]

        ws.cell(row=r, column=COL_PAL).value = f"=SUM(C{r}:H{r})/80"

    for r in range(DATA_ROW_START + len(rows), DATA_ROW_END + 1):
        _style_data_row(ws, r)
        ws.cell(row=r, column=COL_PAL).value = f"=SUM(C{r}:H{r})/80"


def _style_data_row(ws, r: int):
    ws.row_dimensions[r].height = 20.25

    c = ws.cell(row=r, column=COL_DATE)
    c.font = FONT_NORMAL; c.alignment = ALIGN_CENTER
    c.border = Border(top=thin, bottom=thin, left=medium)

    c = ws.cell(row=r, column=COL_REMARK)
    c.font = FONT_NORMAL; c.alignment = ALIGN_CENTER
    c.border = Border(top=thin, bottom=thin, left=double, right=double)

    c = ws.cell(row=r, column=COL_C)
    c.font = FONT_NORMAL; c.alignment = ALIGN_CENTER
    c.border = Border(top=thin, bottom=thin, left=double, right=thin)

    for col in range(COL_C + 1, COL_PAL):
        c = ws.cell(row=r, column=col)
        c.font = FONT_NORMAL; c.alignment = ALIGN_CENTER
        c.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    c = ws.cell(row=r, column=COL_PAL)
    c.font = FONT_NORMAL; c.alignment = ALIGN_CENTER
    c.border = Border(bottom=thin, left=thin, right=double)


def _build_subtotal_row(ws):
    r = TOTAL_ROW
    ws.row_dimensions[r].height = 20.25

    c = ws.cell(row=r, column=COL_DATE)
    c.value = "小計"; c.font = FONT_BOLD; c.alignment = ALIGN_CENTER
    c.border = Border(bottom=medium, left=medium)

    c = ws.cell(row=r, column=COL_REMARK)
    c.font = FONT_BOLD
    c.border = Border(top=thin, bottom=medium, left=double, right=double)

    c = ws.cell(row=r, column=COL_C)
    c.value = f"=SUM(C{DATA_ROW_START}:C{DATA_ROW_END})"
    c.font = FONT_BOLD; c.alignment = ALIGN_CENTER
    c.border = Border(top=thin, bottom=medium, left=double, right=thin)

    for col in range(COL_C + 1, COL_PAL):
        letter = get_column_letter(col)
        c = ws.cell(row=r, column=col)
        c.value = f"=SUM({letter}{DATA_ROW_START}:{letter}{DATA_ROW_END})"
        c.font = FONT_BOLD; c.alignment = ALIGN_CENTER
        c.border = Border(top=thin, bottom=medium, left=thin, right=thin)

    c = ws.cell(row=r, column=COL_PAL)
    c.value = f"=SUM(I{DATA_ROW_START}:I{DATA_ROW_END})"
    c.font = FONT_BOLD; c.alignment = ALIGN_CENTER
    c.border = Border(top=thin, bottom=medium, left=thin, right=double)
