"""
Build the 宏全三廠成品庫存 Excel workbook from accumulated session data.

Column layout (1-indexed):
  A=1  日期
  B=2  所別/備註
  C=3  RGT5F   D=4  RGM5F   E=5  RGG9B   F=6  RGG3B
  G=7  RTT9B   H=8  RTT3B   I=9  RBZ9B   J=10 RBZ3B
  K=11 RBZ5F   L=12 RCB9B   M=13 RCB3B   N=14 GAB5F
  O=15 GBW5F   P=16 GB2M    Q=17 GLW5F   R=18 GL2M
  S=19 GCL5F   T=20 GCL2M   U=21 GHW5F   V=22 GHW2M
  W=23 RYB5F   X=24 KMT5F
  Y=25 0540棧(hidden)  Z=26 0552棧(hidden)  AA=27 0564棧(visible)

Pallet formula (AA col): 9B/3B cols ÷60, 5F/2M cols ÷80
  =INT(SUM(C{r}:D{r},K{r},N{r}:X{r})/80 + SUM(E{r}:J{r},L{r}:M{r})/60)
"""

import io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── 常數 ──────────────────────────────────────────────────────
PRODUCT_CODES = [
    "RGT5F", "RGM5F", "RGG9B", "RGG3B",
    "RTT9B", "RTT3B", "RBZ9B", "RBZ3B",
    "RBZ5F", "RCB9B", "RCB3B", "GAB5F",
    "GBW5F", "GB2M",  "GLW5F", "GL2M",
    "GCL5F", "GCL2M", "GHW5F", "GHW2M",
    "RYB5F", "KMT5F",
]

PRODUCT_LABELS = {
    "RGT5F": "RGT5F\n350ml",
    "RGM5F": "RGM5F\n500ml",
    "RGG9B": "RGG9B\n2000ml",
    "RGG3B": "RGG3B\n600ml",
    "RTT9B": "RTT9B\n2000ml",
    "RTT3B": "RTT3B\n600ml",
    "RBZ9B": "RBZ9B\n2000ml",
    "RBZ3B": "RBZ3B\n600ml",
    "RBZ5F": "RBZ5F\n350ml",
    "RCB9B": "RCB9B\n2000ml",
    "RCB3B": "RCB3B\n600ml",
    "GAB5F": "GAB5F\n350ml",
    "GBW5F": "GBW5F\n350ml",
    "GB2M":  "GB2M\n2000ml",
    "GLW5F": "GLW5F\n350ml",
    "GL2M":  "GL2M\n2000ml",
    "GCL5F": "GCL5F\n350ml",
    "GCL2M": "GCL2M\n2000ml",
    "GHW5F": "GHW5F\n350ml",
    "GHW2M": "GHW2M\n2000ml",
    "RYB5F": "RYB5F\n350ml",
    "KMT5F": "KMT5F\n350ml",
}

# 欄位索引（1-based）
COL_DATE     = 1   # A
COL_REMARK   = 2   # B
COL_FIRST    = 3   # C  (first product)
COL_LAST     = 24  # X  (last product)
COL_PAL_540  = 25  # Y  hidden
COL_PAL_552  = 26  # Z  hidden
COL_PAL_564  = 27  # AA visible

DATA_ROW_START = 4
DATA_ROW_END   = 53
TOTAL_ROW      = 54

FILL_YELLOW = PatternFill("solid", fgColor="FFFACD")
FILL_WHITE  = PatternFill("solid", fgColor="FFFFFF")

FONT_HEADER  = Font(name="Consolas", bold=True, size=9)
FONT_CODE    = Font(name="Consolas", bold=True, size=9, color="CC0000")
FONT_NORMAL  = Font(name="Consolas", size=9)
FONT_TOTAL   = Font(name="Consolas", bold=True, size=9)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")

thin   = Side(style="thin")
medium = Side(style="medium")

def _border(top=thin, bottom=thin, left=thin, right=thin):
    return Border(top=top, bottom=bottom, left=left, right=right)

BORDER_NORMAL = _border()
BORDER_MED_TB = _border(top=medium, bottom=medium)


# ── 公開 API ──────────────────────────────────────────────────

def build_workbook(rows: list) -> bytes:
    """
    rows: list of dicts
      {
        "date":      "2026/05/04",
        "warehouse": "0564",   # used in 所別/備註
        "items":     [{"code": "GLW5F", "qty": 1200}, ...]
      }

    Returns xlsx bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "成品庫存"

    _setup_columns(ws)
    _build_header(ws)
    _fill_data(ws, rows)
    _build_total_row(ws)
    ws.freeze_panes = ws.cell(DATA_ROW_START, COL_FIRST)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 工作表設定 ────────────────────────────────────────────────

def _setup_columns(ws):
    ws.column_dimensions[get_column_letter(COL_DATE)].width    = 11
    ws.column_dimensions[get_column_letter(COL_REMARK)].width  = 10
    for col in range(COL_FIRST, COL_LAST + 1):
        ws.column_dimensions[get_column_letter(col)].width = 7
    ws.column_dimensions[get_column_letter(COL_PAL_540)].width  = 7
    ws.column_dimensions[get_column_letter(COL_PAL_552)].width  = 7
    ws.column_dimensions[get_column_letter(COL_PAL_564)].width  = 8

    ws.column_dimensions[get_column_letter(COL_PAL_540)].hidden = True
    ws.column_dimensions[get_column_letter(COL_PAL_552)].hidden = True

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 36


# ── 表頭 ──────────────────────────────────────────────────────

def _build_header(ws):
    # ── Row 1: 出庫 (C1:X1), 棧板 label at AA1 ──
    # Set borders on each cell first, then merge
    for col in range(COL_FIRST, COL_LAST + 1):
        c = ws.cell(row=1, column=col)
        c.border = _border(top=medium, bottom=thin, left=thin, right=thin)
        c.fill  = FILL_YELLOW
    ws.cell(row=1, column=COL_FIRST).border = _border(top=medium, bottom=thin, left=medium, right=thin)
    ws.cell(row=1, column=COL_LAST).border  = _border(top=medium, bottom=thin, left=thin, right=medium)

    ws.merge_cells(start_row=1, start_column=COL_FIRST,
                   end_row=1,   end_column=COL_LAST)
    c1 = ws.cell(row=1, column=COL_FIRST)
    c1.value     = "出庫"
    c1.font      = FONT_HEADER
    c1.alignment = ALIGN_CENTER
    c1.fill      = FILL_YELLOW

    # AA1 (leave empty, label goes to AA2)
    ws.cell(row=1, column=COL_PAL_564).border = _border(top=medium, bottom=thin, left=thin, right=medium)
    ws.cell(row=1, column=COL_PAL_564).fill   = FILL_YELLOW

    # A1:A2 diagonal / date label — merge rows 1-2 col A
    for r in (1, 2):
        c = ws.cell(row=r, column=COL_DATE)
        c.border = _border(top=medium if r == 1 else thin,
                           bottom=thin if r == 1 else medium,
                           left=medium, right=thin)
        c.fill = FILL_YELLOW
    ws.merge_cells(start_row=1, start_column=COL_DATE,
                   end_row=2,   end_column=COL_DATE)
    ca = ws.cell(row=1, column=COL_DATE)
    ca.value     = "日期"
    ca.font      = FONT_HEADER
    ca.alignment = ALIGN_CENTER
    ca.fill      = FILL_YELLOW

    # ── Row 2: 所別/備註 (B2:B3), 產品 (C2:X2), 棧板 at AA2 ──
    for col in range(COL_FIRST, COL_LAST + 1):
        c = ws.cell(row=2, column=col)
        c.border = _border(top=thin, bottom=medium, left=thin, right=thin)
        c.fill   = FILL_YELLOW
    ws.cell(row=2, column=COL_FIRST).border = _border(top=thin, bottom=medium, left=medium, right=thin)
    ws.cell(row=2, column=COL_LAST).border  = _border(top=thin, bottom=medium, left=thin, right=medium)

    ws.merge_cells(start_row=2, start_column=COL_FIRST,
                   end_row=2,   end_column=COL_LAST)
    c2 = ws.cell(row=2, column=COL_FIRST)
    c2.value     = "產品"
    c2.font      = FONT_HEADER
    c2.alignment = ALIGN_CENTER
    c2.fill      = FILL_YELLOW

    # B2:B3 所別/備註
    for r in (2, 3):
        c = ws.cell(row=r, column=COL_REMARK)
        c.border = _border(top=thin if r == 2 else thin,
                           bottom=thin if r == 2 else medium,
                           left=medium, right=thin)
        c.fill = FILL_YELLOW
    ws.merge_cells(start_row=2, start_column=COL_REMARK,
                   end_row=3,   end_column=COL_REMARK)
    cb = ws.cell(row=2, column=COL_REMARK)
    cb.value     = "所別/備註"
    cb.font      = FONT_HEADER
    cb.alignment = ALIGN_CENTER
    cb.fill      = FILL_YELLOW

    # AA2 棧板
    cpal = ws.cell(row=2, column=COL_PAL_564)
    cpal.value     = "棧板"
    cpal.font      = FONT_HEADER
    cpal.alignment = ALIGN_CENTER
    cpal.fill      = FILL_YELLOW
    cpal.border    = _border(top=thin, bottom=medium, left=thin, right=medium)

    # ── Row 3: 品項 (A3), product codes (C3:X3), 0564(全) (AA3) ──
    ca3 = ws.cell(row=3, column=COL_DATE)
    ca3.value     = "品項"
    ca3.font      = FONT_HEADER
    ca3.alignment = ALIGN_CENTER
    ca3.fill      = FILL_YELLOW
    ca3.border    = _border(top=thin, bottom=medium, left=medium, right=thin)

    for idx, code in enumerate(PRODUCT_CODES):
        col = COL_FIRST + idx
        c   = ws.cell(row=3, column=col)
        c.value     = PRODUCT_LABELS[code]
        c.font      = FONT_CODE
        c.alignment = ALIGN_CENTER
        c.fill      = FILL_YELLOW
        left  = medium if col == COL_FIRST else thin
        right = medium if col == COL_LAST  else thin
        c.border = _border(top=thin, bottom=medium, left=left, right=right)

    cpal3 = ws.cell(row=3, column=COL_PAL_564)
    cpal3.value     = "0564(全)"
    cpal3.font      = FONT_HEADER
    cpal3.alignment = ALIGN_CENTER
    cpal3.fill      = FILL_YELLOW
    cpal3.border    = _border(top=thin, bottom=medium, left=thin, right=medium)


# ── 資料列 ────────────────────────────────────────────────────

def _fill_data(ws, rows: list):
    # Build code→col mapping
    code_to_col = {code: COL_FIRST + i for i, code in enumerate(PRODUCT_CODES)}

    for row_idx, entry in enumerate(rows):
        r = DATA_ROW_START + row_idx
        if r > DATA_ROW_END:
            break

        _style_data_row(ws, r)

        # 日期
        ws.cell(row=r, column=COL_DATE).value = entry.get("date", "")
        ws.cell(row=r, column=COL_DATE).alignment = ALIGN_CENTER

        # 所別/備註 (warehouse)
        ws.cell(row=r, column=COL_REMARK).value = entry.get("warehouse", "")
        ws.cell(row=r, column=COL_REMARK).alignment = ALIGN_CENTER

        # 品項數量
        for item in entry.get("items", []):
            col = code_to_col.get(item["code"])
            if col:
                c = ws.cell(row=r, column=col)
                existing = c.value or 0
                c.value = existing + item["qty"]

        # 棧板公式
        ws.cell(row=r, column=COL_PAL_564).value = (
            f"=INT(SUM(C{r}:D{r},K{r},N{r}:X{r})/80"
            f"+SUM(E{r}:J{r},L{r}:M{r})/60)"
        )
        ws.cell(row=r, column=COL_PAL_564).alignment = ALIGN_CENTER

    # 留空的資料列也要套樣式
    for r in range(DATA_ROW_START + len(rows), DATA_ROW_END + 1):
        _style_data_row(ws, r)
        ws.cell(row=r, column=COL_PAL_564).value = (
            f"=INT(SUM(C{r}:D{r},K{r},N{r}:X{r})/80"
            f"+SUM(E{r}:J{r},L{r}:M{r})/60)"
        )
        ws.cell(row=r, column=COL_PAL_564).alignment = ALIGN_CENTER


def _style_data_row(ws, r: int):
    ws.row_dimensions[r].height = 15
    for col in range(COL_DATE, COL_PAL_564 + 1):
        c = ws.cell(row=r, column=col)
        c.font = FONT_NORMAL
        c.fill = FILL_WHITE
        left  = medium if col in (COL_DATE, COL_REMARK) else thin
        right = medium if col in (COL_REMARK, COL_PAL_564) else thin
        c.border = _border(left=left, right=right)


# ── 合計列 ────────────────────────────────────────────────────

def _build_total_row(ws):
    r = TOTAL_ROW
    ws.row_dimensions[r].height = 16

    ws.cell(row=r, column=COL_DATE).value     = "合計"
    ws.cell(row=r, column=COL_DATE).font      = FONT_TOTAL
    ws.cell(row=r, column=COL_DATE).alignment = ALIGN_CENTER
    ws.cell(row=r, column=COL_DATE).fill      = FILL_YELLOW
    ws.cell(row=r, column=COL_DATE).border    = _border(top=medium, bottom=medium, left=medium, right=thin)

    ws.cell(row=r, column=COL_REMARK).fill   = FILL_YELLOW
    ws.cell(row=r, column=COL_REMARK).border = _border(top=medium, bottom=medium, left=medium, right=thin)

    for col in range(COL_FIRST, COL_LAST + 1):
        letter = get_column_letter(col)
        c = ws.cell(row=r, column=col)
        c.value     = f"=SUM({letter}{DATA_ROW_START}:{letter}{DATA_ROW_END})"
        c.font      = FONT_TOTAL
        c.alignment = ALIGN_CENTER
        c.fill      = FILL_YELLOW
        left  = medium if col == COL_FIRST else thin
        right = medium if col == COL_LAST  else thin
        c.border = _border(top=medium, bottom=medium, left=left, right=right)

    cpal = ws.cell(row=r, column=COL_PAL_564)
    cpal.value = (
        f"=INT(SUM(C{r}:D{r},K{r},N{r}:X{r})/80"
        f"+SUM(E{r}:J{r},L{r}:M{r})/60)"
    )
    cpal.font      = FONT_TOTAL
    cpal.alignment = ALIGN_CENTER
    cpal.fill      = FILL_YELLOW
    cpal.border    = _border(top=medium, bottom=medium, left=thin, right=medium)
