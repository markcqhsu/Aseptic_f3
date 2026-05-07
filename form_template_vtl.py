import os
import sys
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill

WHITE_FILL = PatternFill("solid", fgColor="FFFFFFFF")
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "vtl_template.xlsx")

# 出庫 columns L-CL (col 12-110): product code -> column letter
_CODE_COL_MAP = None

def _get_code_col_map():
    global _CODE_COL_MAP
    if _CODE_COL_MAP is not None:
        return _CODE_COL_MAP
    wb = openpyxl.load_workbook(TEMPLATE_PATH, read_only=True)
    ws = wb.active
    m = {}
    for row_cells in ws.iter_rows(min_row=6, max_row=6, min_col=12, max_col=110):
        for cell in row_cells:
            try:
                v = cell.value
                col_num = cell.column
            except AttributeError:
                continue
            if v and isinstance(v, str):
                code = v.strip().replace("\n", "").replace("\r", "").replace(" ", "").replace("　", "").replace(" ", "")
                if len(code) >= 10:
                    m[code] = get_column_letter(col_num)
    wb.close()
    _CODE_COL_MAP = m
    print(f"[VTL] _get_code_col_map built {len(m)} entries. Sample: {list(m.keys())[:5]}", file=sys.stderr)
    return m


def build_vtl_workbook(rows):
    """
    rows: list of {
        'date':      '2026/05/06',
        'warehouse': '佰事達倉VSR-650188',
        'items':     [{'code': 'JDP0590 1A61', 'qty': 270}, ...]
    }
    Returns xlsx bytes.
    """
    code_col = _get_code_col_map()
    print(f"[VTL] build_vtl_workbook: {len(rows)} rows, code_col size={len(code_col)}", file=sys.stderr)

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # Debug: check template fill types and structure
    sample_cell = ws.cell(row=7, column=1)
    print(f"[VTL] template row7 col1 fill_type={sample_cell.fill.fill_type}", file=sys.stderr)
    try:
        cf_keys = list(ws.conditional_formatting._cf_rules.keys())[:3]
        print(f"[VTL] CF rules: {cf_keys}", file=sys.stderr)
    except Exception as e:
        print(f"[VTL] CF check error: {e}", file=sys.stderr)
    try:
        print(f"[VTL] tables: {list(ws.tables.keys())}", file=sys.stderr)
    except Exception as e:
        print(f"[VTL] tables check error: {e}", file=sys.stderr)

    # Set sheet tab name based on the month of the first row's date
    first_date_str = rows[0].get("date", "") if rows else ""
    try:
        first_dt = datetime.strptime(first_date_str, "%Y/%m/%d")
        ws.title = f"{first_dt.year}.{first_dt.month}月"
    except (ValueError, AttributeError):
        pass

    # Fix header row 6: reset any red/colored font to black
    for cell in ws[6]:
        try:
            if cell.font and cell.font.color and cell.font.color.type == "rgb" \
                    and cell.font.color.rgb not in ("00000000", "FF000000"):
                cell.font = cell.font.copy(color="FF000000")
        except Exception:
            pass

    FIRST_ROW = 7

    # Clear ALL data rows in the template (not just rows with data) for product columns 1-92
    # This removes template background colors from both filled and empty rows
    template_max_row = ws.max_row
    print(f"[VTL] clearing rows 7-{template_max_row} cols 1-92", file=sys.stderr)
    for r in range(FIRST_ROW, template_max_row + 1):
        for c in range(1, 93):
            try:
                ws.cell(row=r, column=c).fill = WHITE_FILL
            except Exception:
                pass

    def safe_set(r, c, v):
        cell = ws.cell(row=r, column=c)
        try:
            cell.value = v
            cell.fill = WHITE_FILL
        except AttributeError:
            pass

    filled_cols = set()

    for i, row in enumerate(rows):
        r = FIRST_ROW + i

        # Column A = date
        date_str = row.get("date", "")
        try:
            dt = datetime.strptime(date_str, "%Y/%m/%d")
        except ValueError:
            dt = None
        safe_set(r, 1, dt)

        # Column K = warehouse
        safe_set(r, 11, row.get("warehouse", ""))

        # Product quantities
        for item in row.get("items", []):
            raw = item.get("code", "")
            normalized = raw.replace(" ", "").replace("　", "").replace(" ", "")
            col_letter = code_col.get(normalized)
            print(f"[VTL] row{i} code={repr(raw)} norm={repr(normalized)} col={col_letter} qty={item.get('qty')}", file=sys.stderr)
            if col_letter:
                cidx = column_index_from_string(col_letter)
                safe_set(r, cidx, item.get("qty", 0))
                filled_cols.add(col_letter)

    print(f"[VTL] filled_cols={filled_cols}", file=sys.stderr)

    # Unhide columns that have data written
    for col_letter in filled_cols:
        ws.column_dimensions[col_letter].hidden = False

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
